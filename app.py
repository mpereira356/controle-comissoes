from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
from io import BytesIO
import re
import os
import unicodedata

from openpyxl import load_workbook

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret-key-for-commission-control'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///comissoes.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Modelo do Banco de Dados
class Comissao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    unid = db.Column(db.Integer, default=1)
    dt_transacao = db.Column(db.Date, nullable=False)
    dt_emissao = db.Column(db.Date, nullable=False)
    pedido = db.Column(db.String(50), nullable=False)
    cod_cli = db.Column(db.String(50))
    cliente = db.Column(db.String(200), nullable=False)
    titulo = db.Column(db.String(50))
    parc = db.Column(db.Integer, default=1)
    ccusto = db.Column(db.String(50))
    dt_vencto = db.Column(db.Date)
    vl_titulo = db.Column(db.Float, default=0.0)
    vl_orig_titulo = db.Column(db.Float, default=0.0)
    comissao_venda = db.Column(db.Float, default=0.0)
    comissao_servico = db.Column(db.Float, default=0.0)
    pedido_erecta = db.Column(db.String(50))
    vendedor = db.Column(db.String(100), nullable=False)
    base_comissao = db.Column(db.Float, default=0.0)
    percentual = db.Column(db.Float, default=10.0)
    vr_comissao = db.Column(db.Float, default=0.0)
    dt_previsao = db.Column(db.Date)
    status = db.Column(db.String(20), default='pendente') # pago, pendente, atrasado
    obs = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'dt_transacao': self.dt_transacao.strftime('%Y-%m-%d'),
            'cliente': self.cliente,
            'pedido': self.pedido,
            'vendedor': self.vendedor,
            'base_comissao': self.base_comissao,
            'percentual': self.percentual,
            'vr_comissao': self.vr_comissao,
            'dt_previsao': self.dt_previsao.strftime('%Y-%m-%d') if self.dt_previsao else None,
            'status': self.status
        }

# Rotas
@app.route('/')
def index():
    vendedor_filtro = request.args.get('vendedor', 'todos')
    status_filtro = request.args.get('status', 'todos')
    cliente_filtro = request.args.get('cliente', '')

    query = Comissao.query

    if vendedor_filtro != 'todos':
        query = query.filter(Comissao.vendedor == vendedor_filtro)
    
    if status_filtro != 'todos':
        query = query.filter(Comissao.status == status_filtro)
    
    if cliente_filtro:
        query = query.filter(Comissao.cliente.ilike(f'%{cliente_filtro}%'))

    comissoes = query.order_by(Comissao.id.asc()).all()
    
    # Calcular totais
    total_valor = sum(c.vl_titulo for c in comissoes)
    total_pago = sum(c.vl_titulo for c in comissoes if c.status == 'pago')
    total_pendente = sum(c.vl_titulo for c in comissoes if c.status == 'pendente')
    total_atrasado = sum(c.vl_titulo for c in comissoes if c.status == 'atrasado')
    
    vendedores = db.session.query(Comissao.vendedor).distinct().all()
    vendedores = [v[0] for v in vendedores]

    return render_template('index.html', 
                           comissoes=comissoes, 
                           total_valor=total_valor,
                           total_pago=total_pago,
                           total_pendente=total_pendente,
                           total_atrasado=total_atrasado,
                           vendedores=vendedores,
                           filtros={'vendedor': vendedor_filtro, 'status': status_filtro, 'cliente': cliente_filtro})

@app.route('/adicionar', methods=['POST'])
def adicionar():
    try:
        # Extrair dados do formulário
        dt_transacao = datetime.strptime(request.form.get('dt_transacao'), '%Y-%m-%d').date()
        dt_emissao = datetime.strptime(request.form.get('dt_emissao'), '%Y-%m-%d').date()
        dt_vencto = datetime.strptime(request.form.get('dt_vencto'), '%Y-%m-%d').date()
        dt_previsao = datetime.strptime(request.form.get('dt_previsao'), '%Y-%m-%d').date()
        
        comissao_venda = float(request.form.get('comissao_venda', 0))
        comissao_servico = float(request.form.get('comissao_servico', 0))
        base_comissao = comissao_venda + comissao_servico
        percentual = float(request.form.get('percentual', 10))
        vr_comissao = base_comissao * (percentual / 100)
        
        # Status automático baseado na data de previsão se não for pago
        status = 'pendente'
        if dt_previsao < datetime.now().date():
            status = 'atrasado'

        nova_comissao = Comissao(
            unid=int(request.form.get('unid', 1)),
            dt_transacao=dt_transacao,
            dt_emissao=dt_emissao,
            pedido=request.form.get('pedido'),
            cod_cli=request.form.get('cod_cli'),
            cliente=request.form.get('cliente'),
            titulo=request.form.get('titulo'),
            parc=int(request.form.get('parc', 1)),
            ccusto=request.form.get('ccusto'),
            dt_vencto=dt_vencto,
            vl_titulo=float(request.form.get('vl_titulo', 0)),
            vl_orig_titulo=float(request.form.get('vl_titulo', 0)),
            comissao_venda=comissao_venda,
            comissao_servico=comissao_servico,
            pedido_erecta=request.form.get('pedido_erecta'),
            vendedor=request.form.get('vendedor'),
            base_comissao=base_comissao,
            percentual=percentual,
            vr_comissao=vr_comissao,
            dt_previsao=dt_previsao,
            status=status,
            obs=request.form.get('obs')
        )
        
        db.session.add(nova_comissao)
        db.session.commit()
        flash('Comissão adicionada com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao adicionar comissão: {str(e)}', 'danger')
    
    return redirect(url_for('index'))

def _normalize_header(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return ''.join(ch for ch in text if ch.isalnum())

def _parse_date(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            continue
    return None

def _parse_float(value):
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace('.', '').replace(',', '.')
    try:
        return float(text)
    except ValueError:
        return 0.0

@app.route('/importar', methods=['POST'])
def importar():
    arquivo = request.files.get('arquivo_excel')
    if not arquivo or arquivo.filename == '':
        flash('Selecione um arquivo Excel para importar.', 'warning')
        return redirect(url_for('index'))

    if not arquivo.filename.lower().endswith('.xlsx'):
        flash('Formato inválido. Envie um arquivo .xlsx.', 'danger')
        return redirect(url_for('index'))

    try:
        wb = load_workbook(filename=BytesIO(arquivo.read()), data_only=True)
        ws = wb.active
        header_row_index = None
        header_candidates = {'pedido', 'cliente', 'vendedor', 'dttransacao', 'dtemissao'}
        for row in ws.iter_rows(min_row=1, max_row=10):
            normalized = [_normalize_header(cell.value) for cell in row]
            if len(header_candidates.intersection(set(normalized))) >= 2:
                header_row_index = row[0].row
                break

        if not header_row_index:
            flash('Não foi possível identificar o cabeçalho da planilha.', 'danger')
            return redirect(url_for('index'))

        headers = {}
        for idx, cell in enumerate(ws[header_row_index], start=1):
            headers[_normalize_header(cell.value)] = idx

        field_map = {
            'unid': ['unid', 'unidade'],
            'dt_transacao': ['dttransacao', 'datatransacao', 'data_transacao'],
            'dt_emissao': ['dtemissao', 'dataemissao', 'data_emissao'],
            'pedido': ['pedido'],
            'cod_cli': ['codcli', 'codcliente', 'codigo_cliente'],
            'cliente': ['cliente'],
            'titulo': ['titulo', 'tituloid'],
            'parc': ['parc', 'parcela'],
            'ccusto': ['ccusto', 'centrocusto'],
            'dt_vencto': ['dtvencto', 'datavencimento', 'vencimento'],
            'vl_titulo': ['vltitulo', 'vltitulor', 'vltitulors', 'valor_titulo', 'valortitulo'],
            'vl_orig_titulo': ['vlorigtitulo', 'vlorigtitr', 'valorigtitulo', 'valor_original_titulo', 'vlorigtitrs'],
            'comissao_venda': ['comissaovenda', 'comissaovenda_rs', 'valorcomissaovendar', 'valorcomissaovendars'],
            'comissao_servico': ['comissaoservico', 'comissaoservicor', 'comissaoservicors', 'valorcomissaoservicor', 'valorcomissaoservicors'],
            'pedido_erecta': ['pedidoerecta', 'pedido_interno', 'pedidointerno', 'dev'],
            'vendedor': ['vendedor'],
            'base_comissao': ['basecomissao', 'basecomissao_rs', 'basecomissoesr', 'basecomissoesrs'],
            'percentual': ['percentual', 'perc', 'comissao'],
            'vr_comissao': ['vrcomissao', 'valorcomissao', 'valor_comissao', 'vrcomissaor', 'vrcomissaors'],
            'dt_previsao': ['dtprevisao', 'dataprevisao', 'previsao_pagamento', 'previsaodepgto'],
            'status': ['status', 'situacao'],
            'obs': ['obs', 'observacao', 'observacoes']
        }

@app.template_filter('dev_short_year')
def dev_short_year(value):
    if not value:
        return ''
    text = str(value)
    return re.sub(r'(?<!\\d)20(\\d{2})(?!\\d)', r'\\1', text)

        col_index = {}
        for field, aliases in field_map.items():
            for alias in aliases:
                if alias in headers:
                    col_index[field] = headers[alias]
                    break

        agregados = {}
        for row_index, row in enumerate(ws.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
            if not any(row):
                continue

            def cell(field):
                idx = col_index.get(field)
                return row[idx - 1] if idx else None

            pedido_erecta = cell('pedido_erecta')
            pedido_erecta = str(pedido_erecta).strip() if pedido_erecta else ''
            chave = _normalize_header(pedido_erecta) if pedido_erecta else f'linha_{row_index}'

            registro = agregados.get(chave)
            if not registro:
                registro = {
                    'unid': int(_parse_float(cell('unid')) or 1),
                    'dt_transacao': _parse_date(cell('dt_transacao')),
                    'dt_emissao': _parse_date(cell('dt_emissao')),
                    'pedido': str(cell('pedido')).strip() if cell('pedido') else '',
                    'cod_cli': str(cell('cod_cli')).strip() if cell('cod_cli') else '',
                    'cliente': str(cell('cliente')).strip() if cell('cliente') else '',
                    'titulo': str(cell('titulo')).strip() if cell('titulo') else '',
                    'parc': int(_parse_float(cell('parc')) or 1),
                    'ccusto': str(cell('ccusto')).strip() if cell('ccusto') else '',
                    'dt_vencto': _parse_date(cell('dt_vencto')),
                    'vl_titulo': _parse_float(cell('vl_titulo')),
                    'vl_orig_titulo': _parse_float(cell('vl_orig_titulo')),
                    'comissao_venda': _parse_float(cell('comissao_venda')),
                    'comissao_servico': _parse_float(cell('comissao_servico')),
                    'pedido_erecta': pedido_erecta,
                    'vendedor': str(cell('vendedor')).strip() if cell('vendedor') else '',
                    'base_comissao': _parse_float(cell('base_comissao')),
                    'percentual': _parse_float(cell('percentual')) or 10.0,
                    'vr_comissao': _parse_float(cell('vr_comissao')),
                    'dt_previsao': _parse_date(cell('dt_previsao')),
                    'status': str(cell('status')).strip().lower() if cell('status') else '',
                    'obs': str(cell('obs')).strip() if cell('obs') else ''
                }
                agregados[chave] = registro
            # Ignorar duplicados do mesmo Pedido Interno (DEV) sem somar valores.
            else:
                continue

        total_importados = 0
        total_atualizados = 0

        for registro in agregados.values():
            if not registro['cliente'] or not registro['vendedor'] or not registro['pedido']:
                continue

            registro['dt_transacao'] = registro['dt_transacao'] or date.today()
            registro['dt_emissao'] = registro['dt_emissao'] or date.today()

            if 0 < registro['percentual'] <= 1:
                registro['percentual'] = registro['percentual'] * 100

            if registro['base_comissao'] == 0.0:
                registro['base_comissao'] = registro['comissao_venda'] + registro['comissao_servico']
            if registro['vr_comissao'] == 0.0:
                registro['vr_comissao'] = registro['base_comissao'] * (registro['percentual'] / 100)

            status = registro['status'] or 'pendente'
            if status != 'pago' and registro['dt_previsao'] and registro['dt_previsao'] < date.today():
                status = 'atrasado'

            existentes = []
            if registro['pedido_erecta']:
                existentes = Comissao.query.filter_by(pedido_erecta=registro['pedido_erecta']).all()

            if existentes:
                existente = existentes[0]
                existente.unid = registro['unid']
                existente.dt_transacao = registro['dt_transacao']
                existente.dt_emissao = registro['dt_emissao']
                existente.pedido = registro['pedido']
                existente.cod_cli = registro['cod_cli']
                existente.cliente = registro['cliente']
                existente.titulo = registro['titulo']
                existente.parc = registro['parc']
                existente.ccusto = registro['ccusto']
                existente.dt_vencto = registro['dt_vencto']
                existente.vl_titulo = registro['vl_titulo']
                existente.vl_orig_titulo = registro['vl_orig_titulo']
                existente.comissao_venda = registro['comissao_venda']
                existente.comissao_servico = registro['comissao_servico']
                existente.pedido_erecta = registro['pedido_erecta']
                existente.vendedor = registro['vendedor']
                existente.base_comissao = registro['base_comissao']
                existente.percentual = registro['percentual']
                existente.vr_comissao = registro['vr_comissao']
                existente.dt_previsao = registro['dt_previsao']
                existente.status = status
                existente.obs = registro['obs']
                for duplicado in existentes[1:]:
                    db.session.delete(duplicado)
                total_atualizados += 1
            else:
                nova_comissao = Comissao(
                    unid=registro['unid'],
                    dt_transacao=registro['dt_transacao'],
                    dt_emissao=registro['dt_emissao'],
                    pedido=registro['pedido'],
                    cod_cli=registro['cod_cli'],
                    cliente=registro['cliente'],
                    titulo=registro['titulo'],
                    parc=registro['parc'],
                    ccusto=registro['ccusto'],
                    dt_vencto=registro['dt_vencto'],
                    vl_titulo=registro['vl_titulo'],
                    vl_orig_titulo=registro['vl_orig_titulo'],
                    comissao_venda=registro['comissao_venda'],
                    comissao_servico=registro['comissao_servico'],
                    pedido_erecta=registro['pedido_erecta'],
                    vendedor=registro['vendedor'],
                    base_comissao=registro['base_comissao'],
                    percentual=registro['percentual'],
                    vr_comissao=registro['vr_comissao'],
                    dt_previsao=registro['dt_previsao'],
                    status=status,
                    obs=registro['obs']
                )
                db.session.add(nova_comissao)
                total_importados += 1

        db.session.commit()
        flash(f'Importação concluída! {total_importados} novos registros e {total_atualizados} atualizados.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao importar Excel: {str(e)}', 'danger')

    return redirect(url_for('index'))

@app.route('/marcar_pago/<int:id>')
def marcar_pago(id):
    comissao = Comissao.query.get_or_404(id)
    comissao.status = 'pago'
    db.session.commit()
    flash('Comissão marcada como paga!', 'success')
    return redirect(url_for('index'))

@app.route('/excluir/<int:id>')
def excluir(id):
    comissao = Comissao.query.get_or_404(id)
    db.session.delete(comissao)
    db.session.commit()
    flash('Registro excluído com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/excluir_multiplos', methods=['POST'])
def excluir_multiplos():
    ids = request.form.getlist('ids')
    if not ids:
        flash('Selecione pelo menos um registro para excluir.', 'warning')
        return redirect(url_for('index'))

    try:
        ids_int = [int(i) for i in ids]
        Comissao.query.filter(Comissao.id.in_(ids_int)).delete(synchronize_session=False)
        db.session.commit()
        flash(f'{len(ids_int)} registro(s) excluído(s) com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir registros: {str(e)}', 'danger')

    return redirect(url_for('index'))

@app.route('/detalhes/<int:id>')
def detalhes(id):
    comissao = Comissao.query.get_or_404(id)
    return jsonify({
        'id': comissao.id,
        'unid': comissao.unid,
        'dt_transacao': comissao.dt_transacao.strftime('%d/%m/%Y'),
        'dt_emissao': comissao.dt_emissao.strftime('%d/%m/%Y'),
        'pedido': comissao.pedido,
        'cod_cli': comissao.cod_cli,
        'cliente': comissao.cliente,
        'titulo': comissao.titulo,
        'parc': comissao.parc,
        'ccusto': comissao.ccusto,
        'dt_vencto': comissao.dt_vencto.strftime('%d/%m/%Y') if comissao.dt_vencto else '',
        'vl_titulo': comissao.vl_titulo,
        'comissao_venda': comissao.comissao_venda,
        'comissao_servico': comissao.comissao_servico,
        'pedido_erecta': comissao.pedido_erecta,
        'vendedor': comissao.vendedor,
        'base_comissao': comissao.base_comissao,
        'percentual': comissao.percentual,
        'vr_comissao': comissao.vr_comissao,
        'dt_previsao': comissao.dt_previsao.strftime('%d/%m/%Y') if comissao.dt_previsao else '',
        'status': comissao.status,
        'obs': comissao.obs
    })

def init_db():
    with app.app_context():
        db.create_all()
        
        # Adicionar dados iniciais se o banco estiver vazio
        if Comissao.query.count() == 0:
            dados_iniciais = [
                {
                    'dt_transacao': '2024-09-10', 'dt_emissao': '2024-08-12', 'pedido': '202025',
                    'cod_cli': '4923', 'cliente': 'REDE MUNICIPAL DR. MARIO GATTI DE URGENCIA, EMERGENCIA E HOSPITALAR',
                    'titulo': '219195', 'parc': 1, 'ccusto': '463109', 'dt_vencto': '2024-09-11',
                    'vl_titulo': 340648.15, 'vl_orig_titulo': 380000, 'comissao_venda': 0,
                    'comissao_servico': 9153.31, 'pedido_erecta': 'DEV-047-2024', 'vendedor': 'Cícero',
                    'base_comissao': 9153.31, 'percentual': 10, 'vr_comissao': 915.33,
                    'dt_previsao': '2024-10-10', 'status': 'pago', 'obs': ''
                },
                {
                    'dt_transacao': '2024-10-10', 'dt_emissao': '2024-09-12', 'pedido': '202025',
                    'cod_cli': '4923', 'cliente': 'REDE MUNICIPAL DR. MARIO GATTI DE URGENCIA, EMERGENCIA E HOSPITALAR',
                    'titulo': '219195', 'parc': 1, 'ccusto': '463109', 'dt_vencto': '2024-10-11',
                    'vl_titulo': 340648.15, 'vl_orig_titulo': 380000, 'comissao_venda': 0,
                    'comissao_servico': 9153.31, 'pedido_erecta': 'DEV-047-2024', 'vendedor': 'Cícero',
                    'base_comissao': 9153.31, 'percentual': 10, 'vr_comissao': 915.33,
                    'dt_previsao': '2024-11-10', 'status': 'pago', 'obs': ''
                },
                {
                    'dt_transacao': '2024-11-11', 'dt_emissao': '2024-10-12', 'pedido': '202025',
                    'cod_cli': '4923', 'cliente': 'REDE MUNICIPAL DR. MARIO GATTI DE URGENCIA, EMERGENCIA E HOSPITALAR',
                    'titulo': '219195', 'parc': 1, 'ccusto': '463109', 'dt_vencto': '2024-11-11',
                    'vl_titulo': 340648.15, 'vl_orig_titulo': 380000, 'comissao_venda': 0,
                    'comissao_servico': 9153.31, 'pedido_erecta': 'DEV-047-2024', 'vendedor': 'Cícero',
                    'base_comissao': 9153.31, 'percentual': 10, 'vr_comissao': 915.33,
                    'dt_previsao': '2024-12-10', 'status': 'pago', 'obs': ''
                },
                {
                    'dt_transacao': '2025-01-10', 'dt_emissao': '2024-10-31', 'pedido': '201331',
                    'cod_cli': '4923', 'cliente': 'REDE MUNICIPAL DR. MARIO GATTI DE URGENCIA, EMERGENCIA E HOSPITALAR',
                    'titulo': '219195', 'parc': 1, 'ccusto': '463109', 'dt_vencto': '2024-11-11',
                    'vl_titulo': 340648.15, 'vl_orig_titulo': 380000, 'comissao_venda': 0,
                    'comissao_servico': 9153.31, 'pedido_erecta': 'DEV-047-2024', 'vendedor': 'Cícero',
                    'base_comissao': 9153.31, 'percentual': 10, 'vr_comissao': 915.33,
                    'dt_previsao': '2025-02-10', 'status': 'pendente', 'obs': ''
                },
                {
                    'dt_transacao': '2025-03-06', 'dt_emissao': '2025-01-29', 'pedido': '206160',
                    'cod_cli': '25563', 'cliente': 'SECRETARIA DE ESTADO DA SAUDE',
                    'titulo': '221991', 'parc': 1, 'ccusto': '463109', 'dt_vencto': '2025-03-10',
                    'vl_titulo': 280000, 'vl_orig_titulo': 280000, 'comissao_venda': 0,
                    'comissao_servico': 6482.05, 'pedido_erecta': 'DEV-001/25', 'vendedor': '50% MPBIOS/CICERO',
                    'base_comissao': 6482.05, 'percentual': 10, 'vr_comissao': 324.10,
                    'dt_previsao': '2025-04-09', 'status': 'pendente', 'obs': 'Comissão dividida 50%'
                },
                {
                    'dt_transacao': '2025-12-11', 'dt_emissao': '2025-08-29', 'pedido': '214341',
                    'cod_cli': '4923', 'cliente': 'REDE MUNICIPAL DR. MARIO GATTI DE URGENCIA, EMERGENCIA E HOSPITALAR',
                    'titulo': '230761', 'parc': 1, 'ccusto': '463109', 'dt_vencto': '2025-11-21',
                    'vl_titulo': 340648.15, 'vl_orig_titulo': 380000, 'comissao_venda': 0,
                    'comissao_servico': 8984.18, 'pedido_erecta': 'DEV-060/25', 'vendedor': 'Cícero',
                    'base_comissao': 8984.18, 'percentual': 10, 'vr_comissao': 898.42,
                    'dt_previsao': '2026-01-08', 'status': 'atrasado', 'obs': 'Pagamento em atraso'
                }
            ]
            
            for d in dados_iniciais:
                c = Comissao(
                    dt_transacao=datetime.strptime(d['dt_transacao'], '%Y-%m-%d').date(),
                    dt_emissao=datetime.strptime(d['dt_emissao'], '%Y-%m-%d').date(),
                    pedido=d['pedido'],
                    cod_cli=d['cod_cli'],
                    cliente=d['cliente'],
                    titulo=d['titulo'],
                    parc=d['parc'],
                    ccusto=d['ccusto'],
                    dt_vencto=datetime.strptime(d['dt_vencto'], '%Y-%m-%d').date(),
                    vl_titulo=d['vl_titulo'],
                    vl_orig_titulo=d['vl_orig_titulo'],
                    comissao_venda=d['comissao_venda'],
                    comissao_servico=d['comissao_servico'],
                    pedido_erecta=d['pedido_erecta'],
                    vendedor=d['vendedor'],
                    base_comissao=d['base_comissao'],
                    percentual=d['percentual'],
                    vr_comissao=d['vr_comissao'],
                    dt_previsao=datetime.strptime(d['dt_previsao'], '%Y-%m-%d').date(),
                    status=d['status'],
                    obs=d['obs']
                )
                db.session.add(c)
            db.session.commit()

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
